import openpyxl
import os
# import random

file_name = os.path.abspath('data.xlsx')
wb = openpyxl.load_workbook(filename=file_name)
sheet = wb['data']

# cell = sheet.cell(row=2, column=2)
# print(cell.value)

rows, columns = sheet.max_row, sheet.max_column

# for row in range(1, rows + 1):
#     for column in range(1, columns + 1):
#         print(sheet.cell(row, column).value, end=" ")
#     print()
# sheet.cell(1, columns + 1).value = "full_name"
# first_name = ""
# last_name = ""
# for row in range(2, rows + 1):
#     for col in [2, 3]:
#         first_name = sheet.cell(row, 2).value
#         last_name = sheet.cell(row, 3).value
#         sheet.cell(row, columns + 1).value = f'{first_name} {last_name}'


# sheet.cell(1, columns + 1).value = "salary"
# for row in range(2, rows + 1):
#     sheet.cell(row, columns + 1).value = random.randrange(start=15000, stop=40000)


sheet.cell(1, columns + 1).value = "avg_salary"
sum_salary = 0

for row in range(2, rows + 1):
    sum_salary += sheet.cell(row, columns).value

sheet.cell(2, columns + 1).value = (sum_salary / (rows - 1))

wb.save(file_name)
print("Excel writing done!!!")
wb.close()
