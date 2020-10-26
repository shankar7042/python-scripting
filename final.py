# file should be like that

# hostname username password database
# abc      shankar  root     abc
import openpyxl
import os
import mysql.connector
from mysql.connector import errorcode

file_name = os.path.abspath('data.xlsx')
wb = openpyxl.load_workbook(filename=file_name)
sheet = wb[wb.sheetnames[0]]

rows, columns = sheet.max_row, sheet.max_column

sheet.cell(1, columns + 1).value = "status"

for row in range(2, rows + 1):
    try:
        hostname = sheet.cell(row, 1).value
        username = sheet.cell(row, 2).value
        password = sheet.cell(row, 3).value
        database = sheet.cell(row, 4).value

        conn = mysql.connector.connect(
            user=username, password=password, host=hostname, database=database)

        sheet.cell(row, columns + 1).value = "Connection Successfull"

    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            sheet.cell(
                row, columns + 1).value = "Something is wrong with your user name or password"
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            sheet.cell(row, columns + 1).value = "Database does not exist"
        else:
            print(err)
    else:
        conn.close()

wb.save(file_name)
print("Excel writing done!!!")
wb.close()
